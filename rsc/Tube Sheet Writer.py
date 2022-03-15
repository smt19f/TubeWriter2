from tkinter import *
from openpyxl import Workbook
import tkinter.messagebox
import math

#Set up window
window = Tk()
window.title("Tube Sheet Writer")
window.geometry("400x400")

#Global flag for error messages
flag = True

#Error message for invalid range
def RangeError():
    tkinter.messagebox.showerror("Error", "Invalid Range")
    flag = False

#Error message for no date
def DateError():
    tkinter.messagebox.showerror("Error", "Date Is Required")
    flag = False

#Converts ranges into lists of ints
def Range2Ints(x):
    result = []
    for part in x.split(',|+'):
        if '-' in part:
            a, b = part.split('-')
            a, b = int(a), int(b)
            if a > b:
                RangeError()
                break
            result.extend(range(a, b+1))
        else:
            a = int(part)
            result.append(a)
    
    return result

#Enters data into excel sheet
def Fill(str, ch, i, total, start, ws):

    array = Range2Ints(str)
    total += len(array)

    for x in array:
            ws.cell(row = i, column = 1, value = start)
            ws.cell(row = i, column = 2, value = ' ' + ch)
            ws.cell(row = i, column = 3, value = (' ' + str(x)) if x < 1000 else x)
            ws.cell(row = i, column = 4, value = ' ' + date)
            i += 1
            start += 1 

#Create excel sheet using entered variables
def Extract():

    #Get variables
    date = entry0.get()
    dog = entry1.get()
    horse = entry2.get()
    bird = entry3.get() 
    double = entry4.get() 
    
    #Start is defaulted to 1
    start = int(entry5.get()) if entry5.get() else 1   

    if not date:
        DateError()

    #Create excel sheet
    wb = Workbook()
    ws = wb.active

    #Fill sheet
    i = 1
    total = 0

    if dog:
        Fill(dog, 'C', i, total, start, ws)

    if horse:
        Fill(horse, 'E', i, total, start, ws)

    if bird:
        Fill(bird, 'A', i, total, start, ws)

    if double:
        Fill(double, 'AS', i, total, start, ws)

    #Export sheet
    if flag:
        wb.save('Tubes.xlsx')

#Display labels and entry fields
dateE = StringVar()
label0 = Label(window, text = "Date (MM/DD): ");   label0.pack()
entry0 = Entry(window, exportselection = 0, textvariable = dateE);   entry0.pack()

dogE = StringVar()
label1 = Label(window, text = "Dogs: ");    label1.pack()
entry1 = Entry(window, exportselection = 0, textvariable = dogE);   entry1.pack()

horseE = StringVar()
label2 = Label(window, text = "Horses: ");   label2.pack()
entry2 = Entry(window, exportselection = 0, textvariable = horseE);   entry2.pack()

birdE = StringVar()
label3 = Label(window, text = "Birds: ");   label3.pack()
entry3 = Entry(window, exportselection = 0, textvariable = birdE);   entry3.pack()

doubleE = StringVar()
label4 = Label(window, text = "Doubles: ");   label4.pack()
entry4 = Entry(window, exportselection = 0, textvariable = doubleE);   entry4.pack()

startE = StringVar()
label5 = Label(window, text = "Start position: ");   label5.pack()
entry5 = Entry(window, exportselection = 0, textvariable = startE);   entry5.pack()

label6 = Label(window)
label6.pack()

submit = Button(window, text = "Submit", command = Extract)
submit.pack()

window.mainloop()